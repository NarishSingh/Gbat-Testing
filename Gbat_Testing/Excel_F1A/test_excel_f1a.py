import json
import openpyxl as xl
import pytest
import requests as rq
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.worksheet import worksheet
from requests import Response
from Excel_F1A.dto.F1A import F1A
from Excel_F1A.dto.Inputs import F1aInput


# region ARRANGE
@pytest.fixture
def get_workbook() -> Workbook:
    """
    Load the excel file to test on
    :return: Workbook object for F1A batch jobs
    """
    return xl.load_workbook(filename=r"C:\MyFiles\VSprojects\Gbat-Testing\Gbat_Testing\Excel_F1A\F1A.xlsx")


@pytest.fixture
def setup_f1a_output_tbl(get_workbook) -> worksheet:
    """
    Get the latest Output Sheet added
    :param: callback for loading F1A workbook
    :return: worksheet object for the latest output table
    """
    f1a: Workbook = get_workbook
    outputs: list[str] = [name for name in f1a.sheetnames if "Output" in name]
    return f1a[outputs[-1]]


@pytest.fixture
def setup_f1a_error_tbl(get_workbook) -> worksheet:
    """
    Get the latest Error sheet added
    :param: callback for loading F1A workbook
    :return: worksheet object for the latest error table/sheet
    """
    f1a: Workbook = get_workbook
    err_sheets: list[str] = [name for name in f1a.sheetnames if "Error" in name]
    return f1a[err_sheets[-1]]


@pytest.fixture
def setup_f1a_addr_range_tbl(get_workbook) -> worksheet:
    """
    Get the latest Address Range sheet added
    :param: callback for loading F1A workbook
    :return: worksheet object for the latest Addr Range table/sheet
    """
    f1a: Workbook = get_workbook
    ar_sheets: list[str] = [name for name in f1a.sheetnames if "Address_Range" in name]
    return f1a[ar_sheets[-1]]


@pytest.fixture
def setup_f1a_complete_bins_tbl(get_workbook) -> worksheet:
    """
    Get the latest Complete BINS sheet added
    :param: callback for loading F1A workbook
    :return: worksheet object for the latest bins table/sheet
    """
    f1a: Workbook = get_workbook
    bin_sheets: list[str] = [name for name in f1a.sheetnames if "Complete_BIN" in name]
    return f1a[bin_sheets[-1]]


@pytest.fixture
def get_expected_output_tbl() -> list[F1A]:
    response_list: list[F1A] = []
    url: str = r"https://geoservice.planning.nyc.gov/geoservice/geoservice.svc/"
    key: str = "qQBPqfVDqJ3ZFJrS"

    # k: id as int, v: input struct with expected input data (will be copied to front of output tbl if user chooses)
    input_tbl: dict = {
        1: F1aInput("1", "120", "Broadway"),
        2: F1aInput("1", "140", "Broadway"),
        3: F1aInput("1", "695", "Park Ave"),
        4: F1aInput("4", "90-15", "queens blvd"),
        5: F1aInput("4", "13118", "liberty ave"),
        6: F1aInput("3", "20", "Fort Greene Pl"),
        7: F1aInput("3", "620", "Atlantic Ave"),
        8: F1aInput("er", "err", "err"),
        9: F1aInput("3", "951", "err3"),
        10: F1aInput("1", "120", "bwa"),
    }

    for k, v in input_tbl.items():
        uri = f"/Function_1A?Borough={v.boro}&ZipCode=&AddressNo={v.addrNo}&StreetName={v.stName}&TPAD=y&BrowseFlag=" \
              f"&HNS=&Unit=&DisplayFormat={False}&Key={key}"
        response: Response = rq.get(url + uri)

        if response.status_code == 200:  # 200 OK -> Geo was able to process request
            f1a_json = json.loads(response.content)
            f1a: F1A = F1A(str(k), v, f1a_json)
            response_list.append(f1a)
        else:
            print("Could not process request")
            print(f"{response.status_code} | {response.content}")

    return response_list


def get_boro_name(boro: str) -> str:
    match boro:
        case "1":
            return "Manhattan"
        case "2":
            return "Bronx"
        case "3":
            return "Brooklyn"
        case "4":
            return "Queens"
        case "5":
            return "Staten Island"
        case _:
            return ""


@pytest.fixture
def get_expected_error_tbl(get_expected_output_tbl) -> list[list[str]]:
    """
    Setup expected errata for an excel batch job querying "boro", "addrNo", "stName", "out_grc", "bbl"
    :return: List of string lists, each representing row data. First row is column headers
    """
    error_responses: list[F1A] = [err for err in get_expected_output_tbl if err.out_grc != "00" and err.out_grc != "01"]

    error_list: list[list[str]] = [
        ["ID", "Function Code", "Borough Code", "Borough Name", "In ZIP Code", "Address Number",
         "Street or Place Name",
         "Unit Input", "GRC", "Error Message", "Similar Name 1", "b7sc 1", "Similar Name 2", "b7sc 2",
         "Similar Name 3",
         "b7sc 3", "Similar Name 4", "b7sc 4", "Similar Name 5", "b7sc 5", "Similar Name 6", "b7sc 6",
         "Similar Name 7",
         "b7sc 7", "Similar Name 8", "b7sc 8", "Similar Name 9", "b7sc 9", "Similar Name 10", "b7sc 10"]
    ]  # col heads

    # error data
    for err in error_responses:
        row: list[str] = [err.ID, err.in_func_code, err.boro, get_boro_name(err.boro), err.in_zip_code, err.addrNo,
                          err.stName, err.unit, err.out_grc, err.out_error_message]
        row.extend(err.similar_names_list)

        error_list.append(row)

    return error_list


# endregion

# region F1A Tests
def test_output_f1a(get_expected_output_tbl, setup_f1a_output_tbl):
    # setting up the expected and ideal/actual values
    expected: list[F1A] = get_expected_output_tbl
    actual: worksheet = setup_f1a_output_tbl
    col_heads: list = list(list(actual.rows)[0])

    # check all cells
    r: int = 0  # row tracker
    for i, row in enumerate(actual.iter_rows()):
        if i == 0:
            continue

        c: int = 0  # col tracker
        cell: Cell
        for cell in row:
            expected_row: F1A = expected[r]
            expected_col_head: Cell = col_heads[c]
            # print(f"{cell.value} | [{r + 2},{c + 1}] {getattr(expected_row, expected_col_head.value)}")  # debug quick print
            assert cell.value == getattr(expected_row, expected_col_head.value)
            c += 1

        r += 1


def test_error_f1a(get_expected_error_tbl, setup_f1a_error_tbl):
    expected: list[list[str]] = get_expected_error_tbl
    actual: worksheet = setup_f1a_error_tbl

    r: int = 0  # row tracker
    for row in actual.rows:
        c: int = 0  # col tracker

        for cell in row:
            # print(f"{cell.value} | [{r + 2},{c + 1}] {expected[r][c]}")  # debug quick print

            # If both are empty and/or all whitespace, test should pass
            test_cell: str = (cell.value or "").strip()
            ideal_cell: str = (expected[r][c] or "").strip()

            assert test_cell == ideal_cell
            c += 1

        r += 1

# endregion
